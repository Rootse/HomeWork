import fdb
import os
from dotenv import load_dotenv
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from datetime import date as dt_date
import pandas as pd
import json
import requests
import shutil
import logging

load_dotenv()


def mm_send_message(url: str, message):
    assert len(url) > 0
    r = requests.post(
        url,
        json=message
    )
    r.raise_for_status()


class Firbird:
    def __init__(self, host: str, user: str, password: str):
        assert len(host) > 0 and len(user) > 0 and len(password) > 0
        self.con = fdb.connect(dsn=host, user=user, password=password, charset='UTF-8')

    def query(self, query: str):
        logging.info("Connect to Firbird...")
        cur = self.con.cursor()
        cur.execute(query)

        return cur


def get_holidays(date_from: str, date_to: str) -> dict:
    logging.info("Get holidays")

    year_from = int(datetime.strptime(date_from, '%Y-%m-%d').strftime('%Y'))
    year_to = int(datetime.strptime(date_to, '%Y-%m-%d').strftime('%Y'))

    holidays = {
        "holidays": [],
        "pre_holidays": []
    }

    assert date_to > date_from
    year = year_from
    while True:
        url = f'https://raw.githubusercontent.com/d10xa/holidays-calendar/master/json/consultant{year}.json'
        r = requests.get(url)
        r = json.loads(r.text)
        holidays['holidays'] += r.get('holidays')
        holidays['pre_holidays'] += r.get('preholidays')

        year += 1

        if year >= year_to:
            break

    return holidays


def get_work_hours_in_month(date_from: str, date_to: str, holidays: dict) -> (int, int):
    date_from = datetime.strptime(date_from, "%Y-%m-%d")
    date_to = datetime.strptime(date_to, "%Y-%m-%d")
    date_today = datetime.today()

    if date_to > date_today:
        date_to = date_today

    assert date_to > date_from
    days = (date_to - date_from).days + 1

    _holidays = len([date for date in holidays['holidays']
                     if date_from <= datetime.strptime(date, "%Y-%m-%d") <= date_to]
                    )

    pre_work_days = len([date for date in holidays['pre_holidays']
                         if date_from <= datetime.strptime(date, "%Y-%m-%d") <= date_to]
                        )

    work_days = int(days) - _holidays

    work_hours = work_days * 9

    assert work_hours > pre_work_days
    work_hours = work_hours - pre_work_days

    return work_days, work_hours


def send_to_mm(ctx: dict):
    url = ctx.get('url')
    username = ctx.get('user')
    message = ctx.get('message')

    assert len(username) > 0
    messages = {
        "username": f"{username}",
        "icon_url": "",
        "text": f"{message}"
    }

    mm_send_message(url, messages)


def main():
    # date_from = "2022-10-01"
    # date_to = "2023-04-30"

    today = datetime.today()
    first_day_of_this_month = datetime(today.year, today.month, 1)

    date_from = datetime(today.year, today.month - 1, 1).strftime("%Y-%m-%d")
    date_to = (first_day_of_this_month - timedelta(days=1)).strftime("%Y-%m-%d")

    holidays = get_holidays(date_from, date_to)

    work_day, work_hours_in_month = get_work_hours_in_month(date_from, date_to, holidays)

    fb_host = os.getenv('DB_HOST')
    fb_user = os.getenv('DB_USER')
    fb_pass = os.getenv('DB_PASS')
    fb = Firbird(fb_host, fb_user, fb_pass)

    data = fb.query(f"""SELECT DISTINCT STAFF.FULL_FIO, M.DATE_PASS, DATEDIFF(MINUTE, (SELECT MIN(table_.TIME_PASS)
                         FROM TABEL_INTERMEDIADATE AS table_
                         WHERE table_.STAFF_ID = M.STAFF_ID AND table_.DATE_PASS = M.DATE_PASS),
                               (SELECT MAX(table_.TIME_PASS)
                         FROM TABEL_INTERMEDIADATE AS table_
                         WHERE table_.STAFF_ID = M.STAFF_ID AND table_.DATE_PASS = M.DATE_PASS))/60 AS hours,
                               MOD(DATEDIFF(MINUTE, (SELECT MIN(table_.TIME_PASS)
                         FROM TABEL_INTERMEDIADATE AS table_
                         WHERE table_.STAFF_ID = M.STAFF_ID AND  table_.DATE_PASS = M.DATE_PASS),
                               (SELECT MAX(table_.TIME_PASS)
                         FROM TABEL_INTERMEDIADATE AS table_
                         WHERE table_.STAFF_ID = M.STAFF_ID AND table_.DATE_PASS = M.DATE_PASS)),60) AS minutes
                         FROM TABEL_INTERMEDIADATE AS M
                         INNER JOIN CONFIGS_TREE ON M.CONFIG_TREE_ID=CONFIGS_TREE.ID_CONFIGS_TREE
                         INNER JOIN STAFF ON M.STAFF_ID = STAFF.ID_STAFF
                         WHERE M.DATE_PASS BETWEEN CAST('{date_from}' AS DATE) AND CAST('{date_to}' AS DATE)
                         ORDER BY STAFF.FULL_FIO, M.DATE_PASS;
                     """)

    df_data = pd.DataFrame(data.fetchall())
    df_data.columns = ['FIO', 'Date', 'Hours', 'Minutes']

    full_data = fb.query(f"""SELECT STAFF.FULL_FIO, M.DATE_PASS, M.TIME_PASS, CONFIGS_TREE.DISPLAY_NAME
                            FROM TABEL_INTERMEDIADATE AS M
                            INNER JOIN CONFIGS_TREE ON M.CONFIG_TREE_ID=CONFIGS_TREE.ID_CONFIGS_TREE
                            INNER JOIN STAFF ON M.STAFF_ID = STAFF.ID_STAFF
                            WHERE M.DATE_PASS BETWEEN CAST('{date_from}' AS DATE) AND CAST('{date_to}' AS DATE)
                            ORDER BY STAFF.FULL_FIO, M.DATE_PASS, M.TIME_PASS;
                         """)

    df_data_full = pd.DataFrame(full_data.fetchall())
    df_data_full.columns = ['FIO', 'Date', 'TimePass', 'DisplayName']

    report = fb.query(f"""SELECT STAFF.FULL_FIO, M.DATE_PASS, DATEDIFF(MINUTE, MIN(M.TIME_PASS), MAX(M.TIME_PASS))
                            FROM TABEL_INTERMEDIADATE AS M
                            INNER JOIN STAFF ON M.STAFF_ID = STAFF.ID_STAFF
                            WHERE STAFF.FULL_FIO NOT SIMILAR TO '\_%|[[:DIGIT:]]%' ESCAPE '\\'
                                  AND ACCESS_BEGIN_DATE IS NOT NULL
                                  AND VALID = 1
                                  AND M.DATE_PASS BETWEEN CAST('{date_from}' AS DATE) AND CAST('{date_to}' AS DATE)
                            GROUP BY M.DATE_PASS, STAFF.FULL_FIO
                            ORDER BY STAFF.FULL_FIO, M.DATE_PASS;
                     """)

    df_report = pd.DataFrame(report.fetchall())
    df_report.columns = ['FIO', 'Date', 'Time']

    logging.info("Add column to report...")
    df_report = df_report.pivot(index='FIO', columns='Date', values='Time')
    df_report.insert(0, "Кол-во_рабочих_часов_в_месяц", work_hours_in_month)
    df_report.insert(1, "Кол-во_рабочих_часов", 0)
    df_report['Кол-во_рабочих_часов'] = df_report.iloc[:, 3:].sum(axis=1)
    df_report.insert(2, "Кол-во_дней_в_офисе", 0)
    df_report['Кол-во_дней_в_офисе'] = df_report.iloc[:, 3:].notnull().astype(int).sum(axis=1)
    df_report.insert(3, "Кол-во_пропущеных_дней", 0)
    df_report['Кол-во_пропущеных_дней'] = df_report.iloc[:, 3:].isnull().sum(axis=1)
    df_report.insert(4, "Кол-во_дней_меньше_9_часов", 0)
    df_report['Кол-во_дней_меньше_9_часов'] = (df_report < 540).iloc[:, 3:].sum(axis=1)

    df_report.sort_values(by=['Кол-во_дней_меньше_9_часов', 'Кол-во_пропущеных_дней'], inplace=True, ascending=False)

    month = (dt_date.today().replace(day=1) - timedelta(days=1)).strftime('%m')
    year = (dt_date.today().replace(day=1) - timedelta(days=1)).strftime('%Y')

    logging.info("Create xlsx...")
    report_name = f'{year}.{month}_Report.xlsx'
    with pd.ExcelWriter(report_name, engine='openpyxl') as writer:
        df_report.to_excel(writer, sheet_name='Report')
        df_data.to_excel(writer, sheet_name='data', index=False)
        df_data_full.to_excel(writer, sheet_name='data_full', index=False)

    logging.info("Write to xlsx...")
    workbook = writer.book
    ws_report = writer.sheets['Report']
    ws_data = writer.sheets['data']
    ws_data_full = writer.sheets['data_full']

    ws_data.column_dimensions['A'].width = 40
    ws_data_full.column_dimensions['A'].width = 40
    ws_report.column_dimensions['A'].width = 45

    ws_data.column_dimensions['B'].width = 20
    ws_data_full.column_dimensions['B'].width = 20
    ws_data_full.column_dimensions['D'].width = 20

    logging.info("Formatting tabel...")
    for col in range(4, ws_report.max_column + 1):
        ws_report.column_dimensions[get_column_letter(col)].width = 15

    for row in range(2, ws_report.max_row + 1):
        for col in range(7, ws_report.max_column + 1):
            if ws_report.cell(row, col).value is None or ws_report.cell(row, col).value == '':
                continue

            # green
            if ws_report.cell(row, col).value >= 540:
                ws_report.cell(row, col).fill = PatternFill(fill_type='solid', start_color='B6E2A1')

            # yellow
            elif ws_report.cell(row, col).value >= 480:
                ws_report.cell(row, col).fill = PatternFill(fill_type='solid', start_color='FEED8C')

            # red
            else:
                ws_report.cell(row, col).fill = PatternFill(fill_type='solid', start_color='F7A4A4')

            ws_report.cell(row, col).value = str(timedelta(minutes=ws_report.cell(row, col).value))
            ws_report.cell(row, col).number_format = 'hh:mm'

        ws_report.cell(row, 3).value = ws_report.cell(row, 3).value / 60
        ws_report.cell(row, 3).number_format = '0.00'

        # ----------- work_house_user -----------
        # red
        if ws_report.cell(row, 3).value < work_hours_in_month * 0.90:
            ws_report.cell(row, 3).fill = PatternFill(fill_type='solid', start_color='F7A4A4')
        # yellow
        elif ws_report.cell(row, 3).value < work_hours_in_month:
            ws_report.cell(row, 3).fill = PatternFill(fill_type='solid', start_color='FEED8C')
        # green
        else:
            ws_report.cell(row, 3).fill = PatternFill(fill_type='solid', start_color='B6E2A1')

        # ----------- days_in_office -----------
        # green
        if ws_report.cell(row, 4).value > work_day * 0.90:
            ws_report.cell(row, 4).fill = PatternFill(fill_type='solid', start_color='B6E2A1')
        # yellow
        elif ws_report.cell(row, 4).value > work_day * 0.75:
            ws_report.cell(row, 4).fill = PatternFill(fill_type='solid', start_color='FEED8C')
        # red
        else:
            ws_report.cell(row, 4).fill = PatternFill(fill_type='solid', start_color='F7A4A4')

        # ----------- missed_days -----------
        # red -
        if ws_report.cell(row, 5).value > work_day * 0.75:
            ws_report.cell(row, 5).fill = PatternFill(fill_type='solid', start_color='F7A4A4')
        # yellow
        elif ws_report.cell(row, 5).value > work_day * 0.50:
            ws_report.cell(row, 5).fill = PatternFill(fill_type='solid', start_color='FEED8C')
        # green
        else:
            ws_report.cell(row, 5).fill = PatternFill(fill_type='solid', start_color='B6E2A1')

    workbook.save(report_name)
    workbook.close()

    path_report = os.getenv('PATH_REPORT')
    path_drive = os.getenv('PATH_DRIVE')

    logging.info("Copy to cloud...")
    src = f"{path_report}/{report_name}"
    dst = f"{path_drive}/{report_name}"
    shutil.copy(src, dst)

    logging.info("Send message...")
    ctx_mm = {
        "url": os.getenv('MM_URL'),
        "user": f"{os.getenv('MM_USER')}",
        "message": f"[Отчет]({os.getenv('DRIVE_URL')})"
    }

    send_to_mm(ctx_mm)


if __name__ == '__main__':
    main()